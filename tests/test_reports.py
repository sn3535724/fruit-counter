from io import BytesIO

from openpyxl import load_workbook

from app.reports import export_excel, export_pdf


def test_export_reports_include_history() -> None:
    history = [
        {
            "timestamp": "04.09.2026 20:00",
            "apple": 1,
            "banana": 2,
            "orange": 3,
            "total": 6,
        }
    ]

    pdf = export_pdf(history).getvalue()
    assert pdf.startswith(b"%PDF")

    workbook = load_workbook(BytesIO(export_excel(history).getvalue()))
    worksheet = workbook["История"]
    assert worksheet.cell(1, 1).value == "Дата и время (МСК)"
    assert worksheet.cell(2, 5).value == 6
